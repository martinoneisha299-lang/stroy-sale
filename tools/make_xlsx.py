#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""catalog.json → КАТАЛОГ-ТАБЛИЦА.xlsx: проверочная таблица для пользователя."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = json.loads(Path("/Users/dm/Desktop/сайт/data/catalog.json").read_text())
TILES = json.loads(Path("/Users/dm/Desktop/сайт/data/tiles.json").read_text())
ROOF_IMG = json.loads(Path("/Users/dm/Desktop/сайт/data/roof_images.json").read_text())
OUT = "/Users/dm/Desktop/сайт/КАТАЛОГ-ТАБЛИЦА.xlsx"

products = DATA["products"]
CAT_RU = {"oblitsovochnyy": "Облицовочный", "obychnyy": "Забутовочный"}
COLL_RU = DATA["collections"]

HEAD = ["ID", "Коллекция", "Категория", "Название на сайте", "Цвет (завод)",
        "Группа цвета", "Фактура", "Формат", "Цена ₽/шт", "Цена ₽/м²",
        "Цены по форматам", "Расход шт/м²", "Фото, шт", "Видео",
        "Проверить", "Заводское название", "Поставщик (на сайт НЕ выводим)"]

wb = Workbook()

# ── Лист «Каталог» ────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Каталог"
ws.append(HEAD)

header_fill = PatternFill("solid", start_color="21201C")
warn_fill = PatternFill("solid", start_color="FFF3C4")
nophoto_fill = PatternFill("solid", start_color="FBE0DC")
thin = Border(bottom=Side(style="thin", color="DDDDDD"))

for c in range(1, len(HEAD) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

for p in sorted(products, key=lambda x: (x["category"], x["collection"] or "я",
                                         x["name"])):
    fp = "; ".join(f"{k}: {v:.2f} ₽" for k, v in p["formats_prices"].items())
    ws.append([
        p["id"],
        COLL_RU.get(p["collection"], "—") if p["collection"] else "—",
        CAT_RU[p["category"]],
        p["name"],
        p["color_raw"] or "",
        p["color_group"],
        p["texture"],
        p["format"],
        p["price"],
        p.get("price_m2"),
        fp,
        p.get("consumption_per_m2"),
        len(p["photos"]),
        "да" if p.get("video") else "",
        "; ".join(p["flags"]),
        p["factory_name"],
        p["supplier"],
    ])

last = ws.max_row
for r in range(2, last + 1):
    for c in range(1, len(HEAD) + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name="Arial", size=10)
        cell.border = thin
    if ws.cell(row=r, column=15).value:                      # флаги
        ws.cell(row=r, column=15).fill = warn_fill
    if ws.cell(row=r, column=13).value == 0:                 # нет фото
        ws.cell(row=r, column=13).fill = nophoto_fill
    ws.cell(row=r, column=9).number_format = "#,##0.00"
    ws.cell(row=r, column=10).number_format = "#,##0.00"

widths = [9, 16, 17, 26, 14, 14, 15, 18, 10, 10, 42, 11, 8, 6, 34, 46, 24]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEAD))}{last}"

# ── Лист «Сводка» ─────────────────────────────────────────────────────────────
sv = wb.create_sheet("Сводка")
sv["A1"] = "Сводка каталога «Строй-Сейл»"
sv["A1"].font = Font(name="Arial", bold=True, size=14)
sv["A2"] = f"Собрано автоматически из папки фото · {DATA['generated']}"
sv["A2"].font = Font(name="Arial", size=9, color="777777")

rows = [
    ("", ""),
    ("Всего товаров", f"=COUNTA(Каталог!A2:A{last})"),
    ("Облицовочный кирпич", f'=COUNTIF(Каталог!C2:C{last},"Облицовочный")'),
    ("Забутовочный", f'=COUNTIF(Каталог!C2:C{last},"Забутовочный")'),
    ("Без фото (докрасим/уточним)", f"=COUNTIF(Каталог!M2:M{last},0)"),
    ("Цена по запросу", f"=COUNTBLANK(Каталог!I2:I{last})"),
    ("С видео", f'=COUNTIF(Каталог!N2:N{last},"да")'),
    ("", ""),
]
r = 3
for label, formula in rows:
    sv.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10)
    if formula:
        c = sv.cell(row=r, column=2, value=formula)
        c.font = Font(name="Arial", size=10, bold=True)
    r += 1

sv.cell(row=r, column=1, value="По коллекциям").font = Font(name="Arial", bold=True, size=11)
r += 1
for slug, name in COLL_RU.items():
    sv.cell(row=r, column=1, value=name).font = Font(name="Arial", size=10)
    sv.cell(row=r, column=2,
            value=f'=COUNTIF(Каталог!B2:B{last},"{name}")').font = Font(name="Arial", size=10, bold=True)
    sv.cell(row=r, column=3,
            value=f'=MINIFS(Каталог!I2:I{last},Каталог!B2:B{last},"{name}")').number_format = "#,##0.00"
    sv.cell(row=r, column=4,
            value=f'=MAXIFS(Каталог!I2:I{last},Каталог!B2:B{last},"{name}")').number_format = "#,##0.00"
    sv.cell(row=r, column=3).font = Font(name="Arial", size=10)
    sv.cell(row=r, column=4).font = Font(name="Arial", size=10)
    r += 1
sv.cell(row=r - len(COLL_RU) - 1, column=3, value="мин. ₽").font = Font(name="Arial", size=9, color="777777")
sv.cell(row=r - len(COLL_RU) - 1, column=4, value="макс. ₽").font = Font(name="Arial", size=9, color="777777")

r += 1
sv.cell(row=r, column=1, value="По цветам (облицовочный)").font = Font(name="Arial", bold=True, size=11)
r += 1
colors = sorted({p["color_group"] for p in products if p["category"] == "oblitsovochnyy"})
for col in colors:
    sv.cell(row=r, column=1, value=col).font = Font(name="Arial", size=10)
    sv.cell(row=r, column=2,
            value=(f'=COUNTIFS(Каталог!F2:F{last},"{col}",'
                   f'Каталог!C2:C{last},"Облицовочный")')).font = Font(name="Arial", size=10, bold=True)
    r += 1

for i, w in enumerate([30, 12, 10, 10], 1):
    sv.column_dimensions[get_column_letter(i)].width = w

# ── Лист «Плитка» ─────────────────────────────────────────────────────────────
tp = wb.create_sheet("Плитка")
TP_HEAD = ["ID", "Форма", "Название на сайте", "Тип окраски", "Цена ₽/м²",
           "Высота", "Поддон", "Морозостойкость", "Фото", "Поставщик (на сайт НЕ выводим)"]
tp.append(TP_HEAD)
for c in range(1, len(TP_HEAD) + 1):
    cell = tp.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

shape_names = {s: m["name"] for s, m in TILES["shapes"].items()}
for p in TILES["products"]:
    tp.append([
        p["id"], shape_names[p["shape"]], p["name"],
        "однотонная" if p["mono"] else "колормикс",
        p["price"],
        p["specs"].get("height", ""),
        p["specs"].get("pallet_m2", ""),
        p["specs"].get("frost", ""),
        "есть" if p["texture"] or p["render"] else "НЕТ",
        "Легион",
    ])
for b in TILES["borders"]:
    tp.append([b["id"], "Бордюры", b["name"], "", b["price"], "", "", "",
               "есть" if b["photo"] else "НЕТ", "Легион"])

tp_last = tp.max_row
for r in range(2, tp_last + 1):
    for c in range(1, len(TP_HEAD) + 1):
        cell = tp.cell(row=r, column=c)
        cell.font = Font(name="Arial", size=10)
        cell.border = thin
    if tp.cell(row=r, column=9).value == "НЕТ":
        tp.cell(row=r, column=9).fill = nophoto_fill
    tp.cell(row=r, column=5).number_format = "#,##0"

for i, w in enumerate([10, 16, 30, 13, 11, 9, 9, 15, 7, 24], 1):
    tp.column_dimensions[get_column_letter(i)].width = w
tp.freeze_panes = "D2"
tp.auto_filter.ref = f"A1:{get_column_letter(len(TP_HEAD))}{tp_last}"

# ── Лист «Кровля» ─────────────────────────────────────────────────────────────
# Товары описаны в tools/build_roof.py (спеки руками) — здесь список для сверки.
import importlib.util as _ilu
_spec = _ilu.spec_from_file_location(
    "build_roof_data", Path(__file__).parent / "build_roof.py")

rf = wb.create_sheet("Кровля")
RF_HEAD = ["ID", "Вид", "Название на сайте", "Коротко", "Цветов", "Фото",
           "Чертёж", "Цена", "Поставщик (на сайт НЕ выводим)"]
rf.append(RF_HEAD)
for c in range(1, len(RF_HEAD) + 1):
    cell = rf.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

# лёгкий парс товаров из build_roof.py без его запуска (он пишет страницы):
import re as _re
_src = (Path(__file__).parent / "build_roof.py").read_text()
_ids = _re.findall(r'id="([a-z0-9-]+)", name="([^"]+)",\s*\n?\s*kind="([^"]+)"', _src)
_meta = dict(_re.findall(
    r'id="((?:mch|prof|sht|sid)-[a-z0-9-]+)".*?meta="([^"]+)"', _src, _re.S))
for pid, name, kind in _ids:
    imgs = ROOF_IMG["products"].get(pid, {})
    rf.append([
        f"krovlya-{pid}", kind, name, _meta.get(pid, ""),
        len(ROOF_IMG["product_colors"].get(pid, [])),
        len(imgs.get("gallery", [])),
        "есть" if imgs.get("scheme") else "нет",
        "по запросу",
        "EURO-Профиль",
    ])
rf_last = rf.max_row
for r in range(2, rf_last + 1):
    for c in range(1, len(RF_HEAD) + 1):
        cell = rf.cell(row=r, column=c)
        cell.font = Font(name="Arial", size=10)
        cell.border = thin
    if rf.cell(row=r, column=6).value == 0 and rf.cell(row=r, column=7).value == "нет":
        rf.cell(row=r, column=6).fill = nophoto_fill
for i, w in enumerate([22, 16, 28, 30, 8, 7, 9, 12, 26], 1):
    rf.column_dimensions[get_column_letter(i)].width = w
rf.freeze_panes = "D2"
rf.auto_filter.ref = f"A1:{get_column_letter(len(RF_HEAD))}{rf_last}"

wb.save(OUT)
print("OK →", OUT)
